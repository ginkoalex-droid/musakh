import io
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.models import Stock, Part, StockMovement, User
from app.auth import get_current_user

router = APIRouter(prefix="/api/export", tags=["export"])


def _style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="1e40af")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


@router.get("/stock")
async def export_stock(
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    result = await db.execute(
        select(Stock, Part).join(Part, Stock.part_id == Part.id).order_by(Part.name)
    )
    rows = result.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Склад"

    headers = ["ID", "Название", "Бренд", "Категория", "Ед.изм", "Место", "Остаток", "Мин.остаток", "Статус"]
    ws.append(headers)
    _style_header(ws)

    for stock, part in rows:
        status = "⚠ Мало" if stock.quantity <= part.min_stock else "OK"
        ws.append([
            part.id, part.name, part.brand, part.category,
            part.unit, part.location, stock.quantity, part.min_stock, status
        ])

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=stock.xlsx"},
    )


@router.get("/movements")
async def export_movements(
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    result = await db.execute(
        select(StockMovement, Part, User)
        .join(Part, StockMovement.part_id == Part.id)
        .join(User, StockMovement.created_by == User.id)
        .order_by(StockMovement.created_at.desc())
    )
    rows = result.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Движения"

    headers = ["ID", "Дата", "Запчасть", "Тип", "Кол-во", "До", "После", "Примечание", "Сотрудник"]
    ws.append(headers)
    _style_header(ws)

    type_labels = {
        "receiving": "Приход",
        "issue": "Выдача",
        "adjustment": "Корректировка",
        "return": "Возврат",
    }

    for mv, part, user in rows:
        ws.append([
            mv.id,
            mv.created_at.strftime("%d.%m.%Y %H:%M"),
            part.name,
            type_labels.get(mv.movement_type.value, mv.movement_type.value),
            mv.quantity,
            mv.quantity_before,
            mv.quantity_after,
            mv.notes or "",
            user.name,
        ])

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=movements.xlsx"},
    )
